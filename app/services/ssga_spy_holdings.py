"""SSGA/SPY holdings workbook company source."""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from app.models.config import AppConfig
from app.services.company_sources import (
    CompanySourceIdentity,
    NormalizedCompanyConstituent,
)


SSGA_SPY_HOLDINGS_URL = (
    "https://www.ssga.com/us/en/individual/library-content/products/"
    "fund-data/etfs/us/holdings-daily-us-en-spy.xlsx"
)
REQUIRED_COLUMNS = ("Name", "Ticker", "Weight")
OPTIONAL_COLUMNS = ("Identifier", "SEDOL", "Sector", "Shares Held", "Local Currency")
NON_COMPANY_MARKERS = (
    "CASH",
    "CURRENCY",
    "DERIVATIVE",
    "FUTURE",
    "SWAP",
    "TREASURY",
    "US DOLLAR",
)


class SSGAWorkbookParseError(ValueError):
    """Raised when the SSGA workbook shape is not usable."""


@dataclass(frozen=True)
class SkippedHoldingRow:
    """A workbook row skipped during provider parsing."""

    row_number: int
    reason: str


@dataclass(frozen=True)
class WorkbookMetadata:
    """Metadata extracted from the top of the holdings workbook."""

    fund_name: str | None = None
    fund_ticker: str | None = None
    holdings_as_of: date | None = None


class SSGASpyHoldingsSource:
    """Company source for the SSGA SPY holdings workbook."""

    identity = CompanySourceIdentity(
        name="ssga_spy_holdings",
        display_name="SSGA SPY holdings",
    )

    def __init__(
        self,
        *,
        workbook_path: str | Path | None = None,
        workbook_url: str = SSGA_SPY_HOLDINGS_URL,
        now: Callable[[], datetime] | None = None,
    ) -> None:
        self.workbook_path = Path(workbook_path) if workbook_path is not None else None
        self.workbook_url = workbook_url
        self.now = now or (lambda: datetime.now(timezone.utc))
        self.last_skipped_rows: list[SkippedHoldingRow] = []
        self.last_metadata = WorkbookMetadata()

    async def fetch(self, config: AppConfig) -> Sequence[Mapping[str, Any]]:
        """Read the workbook and return accepted raw holdings rows."""
        del config
        workbook_bytes = await self._read_workbook()
        rows, metadata, skipped_rows = parse_ssga_spy_holdings_workbook(
            workbook_bytes,
            source_url=self.workbook_url,
            fetched_at=self.now(),
            local_path=str(self.workbook_path) if self.workbook_path else None,
        )
        self.last_metadata = metadata
        self.last_skipped_rows = skipped_rows
        return rows

    def normalize(
        self,
        raw_company: Mapping[str, Any],
        config: AppConfig,
    ) -> NormalizedCompanyConstituent:
        """Normalize one accepted workbook row."""
        del config
        return NormalizedCompanyConstituent.from_source(
            source=self.identity,
            symbol=str(raw_company["Ticker"]),
            name=str(raw_company["Name"]),
            weight=float(raw_company["Weight"]),
            order=int(raw_company["Order"]),
            identifier=optional_text(raw_company.get("Identifier")),
            sedol=optional_text(raw_company.get("SEDOL")),
            sector=optional_text(raw_company.get("Sector")),
            shares_held=optional_float(raw_company.get("Shares Held")),
            local_currency=optional_text(raw_company.get("Local Currency")),
            raw_metadata={
                "source_url": raw_company.get("Source URL"),
                "fetched_at": raw_company.get("Fetched At"),
                "holdings_as_of": raw_company.get("Holdings As Of"),
                "local_path": raw_company.get("Local Path"),
                "fund_name": raw_company.get("Fund Name"),
                "fund_ticker": raw_company.get("Fund Ticker"),
                "provider_row_number": raw_company.get("Provider Row Number"),
                "raw_row": raw_company.get("Raw Row"),
            },
        )

    async def _read_workbook(self) -> bytes:
        if self.workbook_path is not None:
            return self.workbook_path.read_bytes()

        async with httpx.AsyncClient() as client:
            response = await client.get(self.workbook_url)
            response.raise_for_status()
            return response.content


def parse_ssga_spy_holdings_workbook(
    workbook_bytes: bytes,
    *,
    source_url: str,
    fetched_at: datetime,
    local_path: str | None,
) -> tuple[list[dict[str, Any]], WorkbookMetadata, list[SkippedHoldingRow]]:
    """Parse accepted holdings rows from an SSGA/SPY workbook."""
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    if "holdings" not in workbook.sheetnames:
        raise SSGAWorkbookParseError("SSGA workbook must contain a holdings sheet")

    worksheet = workbook["holdings"]
    rows = list(worksheet.iter_rows(values_only=True))
    metadata = extract_metadata(rows)
    header_index, columns = find_header_row(rows)
    skipped_rows: list[SkippedHoldingRow] = []
    accepted_rows: list[dict[str, Any]] = []

    holding_rows = rows[header_index + 1 :]
    for provider_row_number, row in enumerate(holding_rows, start=header_index + 2):
        raw_row = row_to_mapping(row, columns)
        if not any(
            value is not None and str(value).strip() for value in raw_row.values()
        ):
            continue

        reason = skip_reason(raw_row)
        if reason is not None:
            skipped_rows.append(SkippedHoldingRow(provider_row_number, reason))
            continue

        accepted_rows.append(
            {
                **raw_row,
                "Ticker": normalize_ticker(raw_row["Ticker"]),
                "Weight": float(raw_row["Weight"]),
                "Source URL": source_url,
                "Fetched At": fetched_at,
                "Holdings As Of": metadata.holdings_as_of,
                "Local Path": local_path,
                "Fund Name": metadata.fund_name,
                "Fund Ticker": metadata.fund_ticker,
                "Provider Row Number": provider_row_number,
                "Raw Row": raw_row,
            }
        )

    accepted_rows.sort(key=lambda row: (-float(row["Weight"]), str(row["Ticker"])))
    for order, row in enumerate(accepted_rows, start=1):
        row["Order"] = order

    return accepted_rows, metadata, skipped_rows


def extract_metadata(rows: Sequence[Sequence[Any]]) -> WorkbookMetadata:
    """Extract workbook metadata from label/value rows."""
    values: dict[str, Any] = {}
    for row in rows:
        label = normalize_header(row[0]) if row else ""
        if label in {"fund name", "ticker symbol", "holdings"}:
            values[label] = row[1] if len(row) > 1 else None

    return WorkbookMetadata(
        fund_name=optional_text(values.get("fund name")),
        fund_ticker=optional_text(values.get("ticker symbol")),
        holdings_as_of=parse_holdings_as_of(values.get("holdings")),
    )


def find_header_row(rows: Sequence[Sequence[Any]]) -> tuple[int, dict[str, int]]:
    """Find the holdings header row and return header-to-index mapping."""
    for row_index, row in enumerate(rows):
        normalized_to_index = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None and str(value).strip()
        }
        if {"name", "ticker", "weight"}.issubset(normalized_to_index):
            columns = {}
            for column_name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
                normalized_name = normalize_header(column_name)
                if normalized_name in normalized_to_index:
                    columns[column_name] = normalized_to_index[normalized_name]

            missing_columns = [
                column for column in REQUIRED_COLUMNS if column not in columns
            ]
            if missing_columns:
                break
            return row_index, columns

    raise SSGAWorkbookParseError(
        "SSGA holdings sheet is missing required columns: "
        + ", ".join(REQUIRED_COLUMNS)
    )


def row_to_mapping(row: Sequence[Any], columns: Mapping[str, int]) -> dict[str, Any]:
    """Map a worksheet row to known holding columns."""
    return {
        column: row[index] if index < len(row) else None
        for column, index in columns.items()
    }


def skip_reason(row: Mapping[str, Any]) -> str | None:
    """Return why a row should be skipped, or None when accepted."""
    name = optional_text(row.get("Name")) or ""
    ticker = normalize_ticker(row["Ticker"]) if not is_blank(row.get("Ticker")) else ""
    sector = optional_text(row.get("Sector")) or ""
    if is_non_company_holding(name=name, ticker=ticker, sector=sector):
        return "non-company holding"

    for required_column in REQUIRED_COLUMNS:
        if is_blank(row.get(required_column)):
            return f"missing required {required_column}"

    try:
        float(row["Weight"])
    except TypeError, ValueError:
        return "invalid Weight"

    return None


def is_non_company_holding(*, name: str, ticker: str, sector: str) -> bool:
    """Detect obvious non-company/cash/derivative rows."""
    haystack = f"{name} {ticker} {sector}".upper()
    return any(marker in haystack for marker in NON_COMPANY_MARKERS)


def normalize_ticker(value: Any) -> str:
    """Strip and uppercase tickers while preserving punctuation/classes."""
    return str(value).strip().upper()


def normalize_header(value: Any) -> str:
    """Normalize workbook labels and headers for matching."""
    return str(value).strip().rstrip(":").casefold()


def optional_text(value: Any) -> str | None:
    """Return stripped text or None for blank optional fields."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def optional_float(value: Any) -> float | None:
    """Return a float for numeric optional fields."""
    if value is None or str(value).strip() == "":
        return None
    return float(value)


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_holdings_as_of(value: Any) -> date | None:
    """Parse SSGA metadata such as 'As of 09-Jun-2026'."""
    text = optional_text(value)
    if text is None:
        return None
    normalized = text.removeprefix("As of").strip()
    try:
        return datetime.strptime(normalized, "%d-%b-%Y").date()
    except ValueError:
        return None
